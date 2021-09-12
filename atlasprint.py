import lasio
from tkinter import filedialog
from tkinter import *
import tkinter as tk
from PIL import ImageTk, Image
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from win32com import client
from PyPDF2 import PdfFileMerger
from tkinter import messagebox
import time
import threading
import itertools
from numpy import mean
from numpy import std


window = Tk()
window.configure(background='#333')
window.title("ATL PLOT EX. v1.0")
window.bind("<Escape>", lambda e: e.widget.quit())
window.geometry('700x350')
img = ImageTk.PhotoImage(Image.open('C:\AtlasPrint\logo_atlas.png'))
panel = tk.Label(window, image=img, bg='#333')
panel.place(relx=0.5, rely=0.45, anchor=CENTER)
window.iconbitmap('C:\AtlasPrint\icon.ico')
fr = Frame(window)
fr.pack()

firsttext = 'Понимание отдельных принципов освобождает от знания множества фактов\nРичард Фейнман'
secondtext = '+7.916.555.81.95       v. 1.0'
label1 = Label(text=firsttext, fg="gray", bg='#333', anchor=SW, justify='left')
label2 = Label(text=secondtext, fg="gray", bg='#333', anchor=SE)
label3 = Label(text='✎............................................', fg="gray", bg='#333', anchor=SE, justify='right')
label8 = Label(text='', fg="gray", bg='#333', anchor=SE, justify='right')
label1.place(relx=0.01, rely=1, anchor=SW)
label2.place(relx=1, rely=1, anchor=SE)
label3.place(relx=0.8, rely=0.95, anchor=SW)
label8.place(relx=1, rely=0.95, anchor=SE)


def animate():
    for c in itertools.cycle(['✎............................................', '.✎...........................................', '..✎..........................................', '...✎.........................................', '....✎........................................', '.....✎.......................................', '......✎......................................', '.......✎.....................................', '........✎....................................', '.........✎...................................', '..........✎..................................', '...........✎.................................', '............✎................................', '.............✎...............................', '..............✎..............................', '...............✎.............................', '................✎............................', '.................✎...........................', '..................✎..........................', '...................✎.........................', '....................✎........................', '.....................✎.......................', '......................✎......................', '.......................✎.....................', '........................✎....................', '.........................✎...................', '..........................✎..................', '...........................✎.................', '............................✎................', '.............................✎...............', '..............................✎..............', '...............................✎.............', '................................✎............', '.................................✎...........', '..................................✎..........', '...................................✎.........', '....................................✎........', '.....................................✎.......', '......................................✎......', '.......................................✎.....', '........................................✎....', '.........................................✎...', '..........................................✎..', '...........................................✎.', '............................................✎']):
        label3.config(text = (c))
        time.sleep(0.05)
    sys.stdout.write('\rDone!     ')
t = threading.Thread(target=animate)
t.start()


def open_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[(".las", "*.las")])
    label8.config(text='.las loaded')


def save_las():
    label8.config(text='wait...')
    try:

        global data
        with open(file_path, 'r') as f:
            first_data = f.read()
        # replace mnemonics txt file
        new_data2 = first_data.replace('TVD', 'TVDSS')
        new_data3 = new_data2.replace('True Vertical Depth', 'True Vertical Depth Sub Sea')
        new_data4 = new_data3.replace('TRUE VERTICAL DEPTH', 'True Vertical Depth Sub Sea')
        new_data5 = new_data4.replace('VERTICAL DEPTH', 'True Vertical Depth Sub Sea')
        new_data6 = new_data5.replace('VERTICALDEPTH', 'TVDSS')
        new_data7 = new_data6.replace('GAMMA RAY', 'Gamma Ray')
        new_data8 = new_data7.replace('GAMMA', 'GR')
        new_data9 = new_data8.replace('API corrected gamma counts at sensor depth.', 'Gamma Ray')
        new_data10 = new_data9.replace('GAMA', 'GR')
        new_data11 = new_data10.replace('MECHANICAL SPEED', 'Rate Of Penetration')
        new_data12 = new_data11.replace('MECHSPEED', 'ROP')
        new_data13 = new_data12.replace('RATE OF PENERATION', 'Rate Of Penetration')
        new_data14 = new_data13.replace('Rate of penetration at bit depth.', 'Rate Of Penetration')
        new_data15 = new_data14.replace('DEPT', 'DEPTH')
        new_data16 = new_data15.replace('DEPTHH', 'DEPTH')
        new_data17 = new_data16.replace('GRCX', 'GR')
        new_data18 = new_data17.replace('TVDSSSS', 'TVDSS')
        new_data19 = new_data18.replace('MD', 'DEPTH')
        # read data usig lasio
        data = lasio.read(new_data19)
        # data to pandas data frame, rename and sort columns
        df = data.df()
        # sort curves dataframe
        df = df[['GR', 'TVDSS', 'ROP']]
        global tvdss
        tvdss = df['TVDSS']
        #remouve outliers
        upperlimit = df['GR'].mean() + 3 * df['GR'].std()
        lowerlimit = df['GR'].mean() - 3 * df['GR'].std()
        df['GR'] = np.where(df['GR'] > upperlimit, upperlimit, np.where(df['GR'] < lowerlimit, lowerlimit, df['GR']))
        #approximate by moving average (hueta)
        # df['GR'] = df['GR'].rolling(window=3).median()
        #interpolate
        df['GR'] = df['GR'].interpolate(method='cubic')
        #smooth curves
        from scipy.signal import savgol_filter
        df['GR'] = savgol_filter(df['GR'], window_length = 21, polyorder = 13)
        df['ROP'] = savgol_filter(df['ROP'], window_length=21, polyorder=5)
        # pandas to data, write las
        data.set_data(df)
        data.write('C:\\AtlasPrint\\TEMP\\LASFILE.las', version=2.0)
        # read las as txt for rename columns
        with open('C:\\AtlasPrint\\TEMP\\LASFILE.las', 'r') as f:
            renamecolumn = f.read()
        renamecolumn1 = renamecolumn.replace('''DEPTH.M       : DEPTH
    GR   .M       : True Vertical Depth Sub Sea
    TVDSS.API     : Gamma Ray
    ROP  .M.HOUR  : Rate Of Penetration''', '''DEPTH.M    001 : Measured Depth
    GR   .API  --- : Gamma Ray
    TVDSS.M    --- : True Vertical Depth Sub Sea
    ROP  .M/HR --- : Rate Of Penetration''')
        renamecolumn2 = renamecolumn1.replace('''DEPTH.M     : 1  MEASURED DEPTH
    GR   .AAPI  : 2  Gamma Ray
    TVDSS.M     : 3  True Vertical Depth Sub Sea
    ROP  .M/H   : 4  Rate Of Penetration''', '''DEPTH.M    001 : Measured Depth
    GR   .API  --- : Gamma Ray
    TVDSS.M    --- : True Vertical Depth Sub Sea
    ROP  .M/HR --- : Rate Of Penetration''')
        renamecolumn3 = renamecolumn2.replace('''DEPTH.m    : DEPTH
    GR   .m/h  : RATE OF PENETRATION
    TVDSS.API  : Gamma Ray BOREHOLE CORRECTED
    ROP  .m    : True Vertical Depth Sub Sea SUBSEA''', '''DEPTH.M    001 : Measured Depth
        GR   .API  --- : Gamma Ray
        TVDSS.M    --- : True Vertical Depth Sub Sea
        ROP  .M/HR --- : Rate Of Penetration''')
        # read txt lasio and save as las
        renamecolumn4 = lasio.read(renamecolumn3)
        renamecolumn4.write('C:\\AtlasPrint\\TEMP\\LASFILE.las', version=2.0)
        lasname = filedialog.asksaveasfilename(title=u'save file ', filetypes=[(".las", ".las")])
        renamecolumn4.write(lasname, version=2.0)
        label8.config(text='.las saved')

    except Exception as e:
        messagebox.showinfo('Ошибка .las файла', e)
        label8.config(text='ERROR')


def save_md():
    label8.config(text='wait...')
    try:
        label8.config(text='wait...')

        # calc limits
        md_str = data.index
        bottom12 = tvdss.max()
        top12 = tvdss.min()
        bottom12 -= bottom12 % -10
        top12 -= top12 % +10
        bottomMD = md_str.max()
        bottom = md_str.max()
        top = md_str.min()
        bottom -= bottom % -10
        top -= top % +10
        numberdots = len(md_str) + 31
        cm = 1 / 2.54
        ysize = numberdots / 20

        #create fig
        fig, ax = plt.subplots(nrows=1, ncols=2, gridspec_kw={'width_ratios': [1, 2]})
        fig.set_size_inches(21 * cm, ysize * cm)
        fig.subplots_adjust(top=0.88)
        for axes in ax:
            axes.set_ylim(bottom, top - 5)
            depth_major_ticks = np.arange(top - 5, bottom, 5)
            depth_minor_ticks = np.arange(top, bottom, 1)
            axes.set_yticks(depth_major_ticks)
            axes.set_yticks(depth_minor_ticks, minor=True)
            axes.get_xaxis().set_visible(False)
            axes.grid(which='minor', axis='y', alpha=0.5)
            axes.grid(which='major', axis='y', alpha=1)

        ax[1].set_ylabel("Measured Depth [m]", color="Black", fontsize=12, loc='top')

        # track 1 (MD)
        ax_GR = ax[1].twiny()
        ax_GR.set_xlim(0, 150)
        ax_GR.set_xlabel('GR [api]', color="Green", fontsize=12)
        ax_GR.plot(data['GR'], data['DEPTH' or 'DEPT' or 'MD'], color="Green", label='GR [api]', linewidth=1)
        ax_GR.tick_params(axis='x', colors='Green', labeltop=True, labelbottom=True, bottom=True, top=True)
        ax_GR.spines['top'].set_position(('outward', 10))
        ax_GR.spines['top'].set_color('green')
        ax_GR.spines['bottom'].set_position(('outward', 10))
        ax_GR.spines['bottom'].set_color('green')
        major_ticks = np.arange(0, 151, 50)
        minor_ticks = np.arange(0, 151, 10)
        ax_GR.set_xticks(major_ticks)
        ax_GR.set_xticks(minor_ticks, minor = True)
        ax_GR.grid(which='minor', alpha=0.5)
        ax_GR.grid(which='major', alpha=1)

        # track 0 (MD)
        ax_TVDSS = ax[0].twiny()
        ax_TVDSS.set_xlim(bottom12, top12 )
        ax_TVDSS.set_xlabel('TVDSS [m]', color="#ff00ff", fontsize=12)
        ax_TVDSS.plot(data['TVDSS'], data['DEPTH' or 'DEPT' or 'MD'], color="#ff00ff", label='TVDSS [m]', linewidth=1)
        ax_TVDSS.tick_params(axis='x', colors='#ff00ff', labeltop=True, labelbottom=True, bottom=True)
        ax_TVDSS.spines['top'].set_position(('outward', 50))
        ax_TVDSS.spines['top'].set_color('#ff00ff')
        ax_TVDSS.spines['bottom'].set_position(('outward', 50))
        ax_TVDSS.spines['bottom'].set_color('#ff00ff')
        major_ticks1 = np.arange(top12, bottom12+1, (bottom12 - top12)/2)
        minor_ticks1 = np.arange(top12, (bottom12+1 - top12)/2)
        ax_TVDSS.set_xticks(major_ticks1)
        ax_TVDSS.set_xticks(minor_ticks1, minor=True)

        ax_ROP = ax[0].twiny()
        ax_ROP.set_xlim(100, 0)
        ax_ROP.set_xlabel('ROP [m/h]', color="Black", fontsize=12)
        ax_ROP.plot(data['ROP'], data['DEPTH'], color="Black", label='ROP [m/h]', linewidth=1)
        ax_ROP.tick_params(axis='x', colors='Black', labeltop=True, labelbottom=True, bottom=True)
        ax_ROP.spines['top'].set_position(('outward', 10))
        ax_ROP.spines['bottom'].set_position(('outward', 10))
        major_ticks = np.arange(0, 101, 50)
        minor_ticks = np.arange(0, 101, 10)
        ax_ROP.set_xticks(major_ticks)
        ax_ROP.set_xticks(minor_ticks, minor=True)
        ax_ROP.grid(which='minor', alpha=0.5)
        ax_ROP.grid(which='major', alpha=1)

        fig.tight_layout()
        fig.savefig('C:\\AtlasPrint\\TEMP\\GR_PLOT_PDF_md.pdf')

        # write changes to xlsx header
        path = "C:\\AtlasPrint\\Header MD.xlsx"
        wb_obj = openpyxl.load_workbook(path.strip())
        sheet_obj = wb_obj.active
        cellThatIsToBeChanged = sheet_obj.cell(row=43, column=2)
        cellThatIsToBeChanged.value = bottomMD
        wb_obj.save('C:\\AtlasPrint\\Header MD.xlsx')
        label8.config(text='wait...')
        # save xlsx to pdf

        excel = client.Dispatch("Excel.Application")
        sheets = excel.Workbooks.Open('C:\\AtlasPrint\\Header MD.xlsx')
        work_sheets = sheets.Worksheets[0]
        work_sheets.ExportAsFixedFormat(0, 'C:\\AtlasPrint\\TEMP\\Header MD.pdf')
        sheets.Close(True)

        # merge pdf MD
        pdfs = ['C:\\AtlasPrint\\TEMP\\Header MD.pdf', 'C:\\AtlasPrint\\TEMP\\GR_PLOT_PDF_md.pdf']
        merger = PdfFileMerger()
        for pdf in pdfs:
            merger.append(pdf)
        pdfmdname = filedialog.asksaveasfilename(title=u'save file ', filetypes=[(".pdf", ".pdf")])
        merger.write(pdfmdname)
        merger.close()
        label8.config(text='md saved')
    except Exception as e:
        messagebox.showinfo('Ошибка MD.pdf', e)
        label8.config(text='ERROR')


def save_tvdss():
    label8.config(text='wait...')
    try:
        # create fig TVDSS
        bottom1 = tvdss.max()
        top1 = tvdss.min()
        bottom1 -= bottom1 % -10
        top1 -= top1 % +10
        numberdots1 = len(tvdss)
        ysize1 = numberdots1 / 40
        cm = 1 / 2.54

        fig1, ax = plt.subplots(nrows=1, ncols=2, gridspec_kw={'width_ratios': [1, 2]})
        fig1.set_size_inches(21 * cm, ysize1 * cm)
        fig1.suptitle("", fontsize=20)
        fig1.subplots_adjust(top=0.88)

        for axes in ax:
            axes.set_ylim(bottom1, top1)
            depth_major_ticks = np.arange(top1 - 2, bottom1 + 2, 5)
            depth_minor_ticks = np.arange(top1 - 2, bottom1 + 2, 1)
            axes.set_yticks(depth_major_ticks)
            axes.set_yticks(depth_minor_ticks, minor=True)
            axes.set_yticks(depth_minor_ticks, minor=True)
            axes.get_xaxis().set_visible(False)
            axes.grid(which='minor', axis='y', alpha=0.5)
            axes.grid(which='major', axis='y', alpha=1)

        ax[1].set_ylabel("True Vertical Depth Sub Sea [m]", color="Black", fontsize=12, loc='top')
        # track 1 (TVDSS)
        ax_GR = ax[1].twiny()
        ax_GR.set_xlim(0, 150)
        ax_GR.set_xlabel('GR [api]', color="Green", fontsize=12)
        ax_GR.plot(data['GR'], data['TVDSS'], color="Green", label='GR [api]', linewidth=1)
        ax_GR.tick_params(axis='x', colors='Green', labeltop=True, labelbottom=True, bottom=True, top=True)
        ax_GR.spines['top'].set_position(('outward', 10))
        ax_GR.spines['top'].set_color('green')
        ax_GR.spines['bottom'].set_position(('outward', 10))
        ax_GR.spines['bottom'].set_color('green')
        major_ticks = np.arange(0, 151, 50)
        minor_ticks = np.arange(0, 151, 10)
        ax_GR.set_xticks(major_ticks)
        ax_GR.set_xticks(minor_ticks, minor = True)
        ax_GR.grid(which='minor', alpha=0.5)
        ax_GR.grid(which='major', alpha=1)
        # track 0 (TVDSS)
        ax_ROP = ax[0].twiny()
        ax_ROP.set_xlim(100, 0)
        ax_ROP.set_xlabel('ROP [m/h]', color="Black", fontsize=12)
        ax_ROP.plot(data['ROP'], data['TVDSS'], color="Black", label='ROP [m/h]', linewidth=1)
        ax_ROP.tick_params(axis='x', colors='Black', labeltop=True, labelbottom=True, bottom=True)
        ax_ROP.spines['top'].set_position(('outward', 10))
        ax_ROP.spines['bottom'].set_position(('outward', 10))
        major_ticks = np.arange(0, 101, 50)
        minor_ticks = np.arange(0, 101, 10)
        ax_ROP.set_xticks(major_ticks)
        ax_ROP.set_xticks(minor_ticks, minor=True)
        ax_ROP.grid(which='minor', alpha=0.5)
        ax_ROP.grid(which='major', alpha=1)
        fig1.tight_layout()

        fig1.savefig('C:\\AtlasPrint\\TEMP\\GR_PLOT_PDF_tvdss.pdf')

        # write chsnges to xlsx
        path = "C:\\AtlasPrint\\Header TVDSS.xlsx"
        wb_obj = openpyxl.load_workbook(path.strip())
        sheet_obj = wb_obj.active
        cellThatIsToBeChanged = sheet_obj.cell(row=43, column=2)
        cellThatIsToBeChanged.value = bottom1
        wb_obj.save('C:\\AtlasPrint\\Header TVDSS.xlsx')

        # convert xslasx to pdf
        excel = client.Dispatch("Excel.Application")
        sheets1 = excel.Workbooks.Open('C:\\AtlasPrint\\Header TVDSS.xlsx')
        work_sheets1 = sheets1.Worksheets[0]
        work_sheets1.ExportAsFixedFormat(0, 'C:\\AtlasPrint\\TEMP\\Header TVDSS.pdf')
        sheets1.Close(True)

        # merge pdf TVDSS
        from PyPDF2 import PdfFileMerger
        pdfs = ['C:\\AtlasPrint\\TEMP\\Header TVDSS.pdf', 'C:\\AtlasPrint\\TEMP\\GR_PLOT_PDF_tvdss.pdf']
        merger = PdfFileMerger()
        for pdf in pdfs:
            merger.append(pdf)
        pdftvdssname = filedialog.asksaveasfilename(title=u'save file ', filetypes=[(".pdf", ".pdf")])
        merger.write(pdftvdssname)
        merger.close()
        label8.config(text='')

    except Exception as e:
        messagebox.showinfo('Ошибка TVDSS.pdf', e)
        label8.config(text='ERROR')


bt1 = Button(fr, width=20, text='OPEN .las', relief=RAISED, bd=6, bg='#333', fg='white', command = open_file)
bt2 = Button(fr, width=15, text='save .las', relief=RAISED, bd=6, bg='#333', fg='white', command = save_las)
bt3 = Button(fr, width=15, text='print md', relief=RAISED, bd=6, bg='#333', fg='white', command = save_md)
bt4 = Button(fr, width=15, text='print tvdss', relief=RAISED, bd=6, bg='#333', fg='white', command = save_tvdss)
bt1.pack(side='left', )
bt2.pack(side='left', padx=0)
bt3.pack(side='left', padx=0)
bt4.pack(side='left', padx=0)

window.mainloop()
